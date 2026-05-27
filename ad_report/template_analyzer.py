from __future__ import annotations

import json
import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .dictionary import StandardDictionary, normalize_token


TITLE_HINTS = ("top", "데이터", "성과", "요약", "분석", "키워드", "매체", "기기", "소재", "캠페인")
MAX_SCAN_ROWS = 500


def list_workbook_sheets(workbook_bytes: bytes) -> list[str]:
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)
    return workbook.sheetnames


def analyze_template(
    workbook_bytes: bytes,
    dictionary: StandardDictionary,
    sheet_names: list[str] | None = None,
) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    definitions: list[dict[str, Any]] = []
    selected_sheets = set(sheet_names or workbook.sheetnames)

    for sheet in workbook.worksheets:
        if sheet.title not in selected_sheets:
            continue
        max_row = min(sheet.max_row or 1, MAX_SCAN_ROWS)
        for row in range(1, max_row + 1):
            title = _row_text(sheet, row)
            header_row = _find_header_row(sheet, row, dictionary)
            if not header_row:
                continue
            title_cell = _first_non_empty_cell(sheet, row)
            header_map = _header_map(sheet, header_row, dictionary)
            if not _looks_like_new_table(definitions, sheet.title, header_row):
                continue
            definition = _definition_from_context(
                sheet=sheet,
                sheet_name=sheet.title,
                title=title,
                title_cell=title_cell,
                header_row=header_row,
                header_map=header_map,
                dictionary=dictionary,
            )
            definitions.append(definition)

    return definitions


def analyze_with_gemini(
    workbook_bytes: bytes,
    dictionary: StandardDictionary,
    schema: dict[str, Any],
    api_key: str,
    model: str,
    sheet_names: list[str] | None = None,
) -> list[dict[str, Any]]:
    try:
        from google import genai
    except Exception:
        return analyze_template(workbook_bytes, dictionary, sheet_names)

    client = genai.Client(api_key=api_key)
    definitions: list[dict[str, Any]] = []
    for sheet_outline in extract_workbook_outline(workbook_bytes, dictionary, sheet_names):
        prompt = _gemini_prompt(dictionary, schema, [sheet_outline])
        try:
            response = client.models.generate_content(
                model=model,
                contents=prompt,
                config={"temperature": 0, "response_mime_type": "application/json"},
            )
            content = response.text or "{}"
            parsed = json.loads(content)
            if isinstance(parsed, dict):
                definitions.extend(parsed.get("tables", parsed.get("definitions", [])))
            elif isinstance(parsed, list):
                definitions.extend(parsed)
        except Exception:
            continue
    if definitions:
        return definitions
    return analyze_template(workbook_bytes, dictionary, sheet_names)


def _gemini_prompt(dictionary: StandardDictionary, schema: dict[str, Any], outline: list[dict[str, Any]]) -> str:
    return (
        "엑셀 광고 보고서 템플릿의 표 정의를 추출하세요. "
        "반드시 JSON 객체만 반환하세요. 최상위 키는 definitions 또는 tables 입니다. "
        "각 표는 id, name, group_by, sort, limit, metrics, location, metadata를 포함합니다. "
        "헤더 아래에 합계/총계 행이 있으면 total_row.row에 행 번호를 넣고, "
        "일평균/평균 행이 있으면 average_row.row에 행 번호를 넣고, "
        "전월 비교/전주 비교/증감 행이 있으면 compare_row.row에 행 번호를 넣으세요. "
        "수식 셀이 있으면 formula_cells에 셀 주소/행/컬럼/수식을 넣고, "
        "GETPIVOTDATA 수식이면 source_type='pivot_formula', formula_policy='preserve'로 설정하세요. "
        "location.data_start_row는 실제 상세 데이터가 시작되는 첫 행이어야 합니다.\n\n"
        f"표준 차원: {dictionary.dimension_names}\n"
        f"표준 지표: {dictionary.metric_names}\n"
        f"JSON Schema: {json.dumps(schema, ensure_ascii=False)}\n"
        f"Workbook outline: {json.dumps(outline, ensure_ascii=False)}"
    )


def extract_workbook_outline(
    workbook_bytes: bytes,
    dictionary: StandardDictionary,
    sheet_names: list[str] | None = None,
) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False, read_only=True)
    outline = []
    selected_sheets = set(sheet_names or workbook.sheetnames)
    for sheet in workbook.worksheets:
        if sheet.title not in selected_sheets:
            continue
        rows = []
        max_row = min(sheet.max_row or 1, 80)
        for row_idx, row in enumerate(sheet.iter_rows(max_row=max_row, values_only=False), start=1):
            values = [cell.value for cell in row[:20]]
            if any(value is not None and str(value).strip() for value in values):
                matches = [dictionary.match(value) for value in values]
                formulas = [
                    {
                        "column": get_column_letter(cell.column),
                        "formula": str(cell.value),
                    }
                    for cell in row[:20]
                    if _is_formula(cell.value)
                ]
                rows.append(
                    {
                        "row": row_idx,
                        "values": ["" if value is None else str(value) for value in values],
                        "matches": matches,
                        "formulas": formulas[:5],
                    }
                )
        outline.append({"sheet": sheet.title, "rows": rows})
    return outline


def _definition_from_context(
    sheet: Any,
    sheet_name: str,
    title: str,
    title_cell: str,
    header_row: int,
    header_map: dict[str, str],
    dictionary: StandardDictionary,
) -> dict[str, Any]:
    dimensions = [value for value in header_map.values() if dictionary.is_dimension(value)]
    metrics = [value for value in header_map.values() if dictionary.is_metric(value)]
    title_matches = _matches_in_text(title, dictionary)

    group_by = [match for match in title_matches if dictionary.is_dimension(match)]
    if not group_by and dimensions:
        group_by = [dimensions[0]]
    if not group_by:
        group_by = ["매체"]

    limit = _infer_limit(title)
    sort_by = _infer_sort(title, metrics, group_by, dictionary)
    label_col = _label_col(header_map, group_by)
    metric_columns = {
        col: name
        for col, name in header_map.items()
        if dictionary.is_metric(name) and col != label_col
    }
    if not metrics:
        metrics = list(metric_columns.values())

    safe_name = title.strip() or f"{sheet_name} {header_row}행 표"
    summary_rows = _detect_summary_rows(sheet, header_row, label_col)
    fixed_summary_rows = [spec["row"] for spec in summary_rows.values() if spec.get("row")]
    data_start = max([header_row + 1, *[row + 1 for row in fixed_summary_rows]])
    data_end = data_start + (limit - 1 if limit else 19)
    formula_cells = _detect_formula_cells(sheet, header_row + 1, data_end, [label_col, *metric_columns.keys()])

    definition = {
        "id": _slugify(f"{sheet_name}_{safe_name}_{header_row}"),
        "name": safe_name,
        "group_by": group_by,
        "sort": {"by": sort_by, "order": "desc"},
        "metrics": metrics,
        "location": {
            "sheet": sheet_name,
            "title_cell": title_cell,
            "header_row": header_row,
            "label_col": label_col,
            "data_start_row": data_start,
            "data_end_row": data_end,
            "columns": metric_columns,
        },
        "metadata": {"created_by": "ai", "ai_confidence": 0.72, "user_verified": False},
    }
    if formula_cells:
        definition["formula_cells"] = formula_cells
        definition["formula_policy"] = "preserve"
        definition["source_type"] = "pivot_formula" if any("GETPIVOTDATA" in cell["formula"].upper() for cell in formula_cells) else "formula_table"
    if "total_row" in summary_rows:
        definition["total_row"] = summary_rows["total_row"]
    if "average_row" in summary_rows:
        definition["average_row"] = summary_rows["average_row"]
    elif not limit:
        definition["total_row"] = {"enabled": True, "position": "top", "label": "합계"}
    if "compare_row" in summary_rows:
        definition["compare_row"] = summary_rows["compare_row"]
    if limit:
        definition["limit"] = limit
    return definition


def _row_text(sheet: Any, row: int) -> str:
    values = [
        str(cell.value).strip()
        for cell in sheet[row]
        if cell.value is not None and str(cell.value).strip()
    ]
    return " ".join(values)


def _first_non_empty_cell(sheet: Any, row: int) -> str:
    for cell in sheet[row]:
        if cell.value is not None and str(cell.value).strip():
            return cell.coordinate
    return f"A{row}"


def _find_header_row(sheet: Any, row: int, dictionary: StandardDictionary) -> int | None:
    title = _row_text(sheet, row)
    if not title:
        return row if _is_header_row(sheet, row, dictionary) else None
    if title and not any(hint in normalize_token(title) for hint in TITLE_HINTS):
        matches = _matches_in_text(title, dictionary)
        if not matches:
            return None

    for candidate in range(row, min(row + 5, sheet.max_row or row) + 1):
        if _is_header_row(sheet, candidate, dictionary):
            return candidate
    return None


def _is_header_row(sheet: Any, row: int, dictionary: StandardDictionary) -> bool:
    header_map = _header_map(sheet, row, dictionary)
    metric_count = sum(1 for value in header_map.values() if dictionary.is_metric(value))
    dimension_count = sum(1 for value in header_map.values() if dictionary.is_dimension(value))
    return (metric_count >= 1 and dimension_count >= 1) or metric_count >= 2


def _header_map(sheet: Any, row: int, dictionary: StandardDictionary) -> dict[str, str]:
    mapping = {}
    for cell in sheet[row]:
        match = dictionary.match(cell.value)
        if match:
            mapping[get_column_letter(cell.column)] = match
    return mapping


def _matches_in_text(text: str, dictionary: StandardDictionary) -> list[str]:
    matches = []
    normalized = normalize_token(text)
    for name in [*dictionary.dimension_names, *dictionary.metric_names]:
        tokens = [name, *dictionary.dimensions.get(name, {}).get("synonyms", []), *dictionary.metrics.get(name, {}).get("synonyms", [])]
        if any(normalize_token(token) in normalized for token in tokens):
            matches.append(name)
    return matches


def _infer_limit(title: str) -> int | None:
    match = re.search(r"(?:top|TOP|상위)\s*([0-9]+)", title)
    return int(match.group(1)) if match else None


def _infer_sort(title: str, metrics: list[str], group_by: list[str], dictionary: StandardDictionary) -> str:
    title_matches = [match for match in _matches_in_text(title, dictionary) if dictionary.is_metric(match)]
    if title_matches:
        return title_matches[0]
    if metrics:
        return metrics[0]
    return group_by[0]


def _label_col(header_map: dict[str, str], group_by: list[str]) -> str:
    for col, name in header_map.items():
        if name in group_by:
            return col
    return next(iter(header_map.keys()), "A")


def _detect_summary_rows(sheet: Any, header_row: int, label_col: str) -> dict[str, dict[str, Any]]:
    summary_rows: dict[str, dict[str, Any]] = {}
    label_col_idx = 1
    for cell in sheet[header_row]:
        if get_column_letter(cell.column) == label_col:
            label_col_idx = cell.column
            break
    for row in range(header_row + 1, min(header_row + 5, sheet.max_row or header_row) + 1):
        value = sheet.cell(row, label_col_idx).value
        text = "" if value is None else str(value).strip()
        normalized = normalize_token(text)
        if "평균" in normalized:
            summary_rows["average_row"] = {
                "enabled": True,
                "row": row,
                "label": text or "일평균",
                "mode": "daily_average",
            }
        elif "합계" in normalized or "총계" in normalized:
            summary_rows["total_row"] = {"enabled": True, "row": row, "label": text or "합계"}
        elif "비교" in normalized or "전월" in normalized or "전주" in normalized or "증감" in normalized:
            summary_rows["compare_row"] = {
                "enabled": True,
                "row": row,
                "label": text or "전월 비교",
                "mode": "previous_row",
            }
    return summary_rows


def _detect_formula_cells(sheet: Any, start_row: int, end_row: int, columns: list[str]) -> list[dict[str, Any]]:
    formula_cells = []
    for row in range(start_row, min(end_row, sheet.max_row or end_row) + 1):
        for column in columns:
            cell = sheet[f"{column}{row}"]
            if _is_formula(cell.value):
                formula_cells.append(
                    {
                        "cell": cell.coordinate,
                        "row": row,
                        "column": column,
                        "formula": _truncate_formula(str(cell.value)),
                    }
                )
    return formula_cells


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _truncate_formula(formula: str, max_length: int = 240) -> str:
    return formula if len(formula) <= max_length else f"{formula[:max_length]}..."


def _looks_like_new_table(definitions: list[dict[str, Any]], sheet: str, header_row: int) -> bool:
    for definition in definitions:
        location = definition.get("location", {})
        if location.get("sheet") == sheet and abs(location.get("header_row", 0) - header_row) < 3:
            return False
    return True


def _slugify(value: str) -> str:
    slug = re.sub(r"[^0-9a-zA-Z가-힣]+", "_", value).strip("_").lower()
    return slug or "table_definition"
