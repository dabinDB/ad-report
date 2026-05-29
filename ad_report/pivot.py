from __future__ import annotations

from io import BytesIO
import re
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl.utils import get_column_letter, range_boundaries

from .dictionary import StandardDictionary


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
NS = {"main": MAIN_NS, "rel": REL_NS, "pkgrel": PACKAGE_REL_NS}
ET.register_namespace("", MAIN_NS)


def detect_pivot_sources(workbook_bytes: bytes) -> list[dict[str, Any]]:
    sources = []
    with ZipFile(BytesIO(workbook_bytes), "r") as archive:
        pivot_defs = [
            name
            for name in archive.namelist()
            if name.startswith("xl/pivotCache/") and name.endswith(".xml")
        ]
        for index, path in enumerate(sorted(pivot_defs), start=1):
            root = ET.fromstring(archive.read(path))
            worksheet_source = root.find(".//main:worksheetSource", NS)
            if worksheet_source is None:
                continue
            ref = worksheet_source.attrib.get("ref")
            sheet = worksheet_source.attrib.get("sheet")
            source_name = worksheet_source.attrib.get("name")
            effective_ref = _effective_source_ref(archive, sheet, ref)
            sources.append(
                {
                    "id": f"pivot_cache_{index}",
                    "cache_definition": path,
                    "sheet": sheet,
                    "ref": effective_ref,
                    "original_ref": ref,
                    "name": source_name,
                    "display": _source_display(sheet, effective_ref, source_name),
                    "row_count": _source_row_count(effective_ref),
                    "refresh_on_load": root.attrib.get("refreshOnLoad") == "1",
                }
            )
    return sources


def update_pivot_source_data(
    workbook_bytes: bytes,
    source_df: pd.DataFrame,
    pivot_source: dict[str, Any],
    mode: str,
    dictionary: StandardDictionary,
) -> bytes:
    sheet_name = pivot_source.get("sheet")
    if not sheet_name:
        raise ValueError(f"피벗 소스 시트를 찾을 수 없습니다: {sheet_name}")
    ref = pivot_source.get("ref")
    if not ref:
        raise ValueError("피벗 소스 범위가 없어 대용량 직접 갱신을 할 수 없습니다.")

    min_col, header_row, max_col, max_row = range_boundaries(ref)
    input_buffer = BytesIO(workbook_bytes)
    output_buffer = BytesIO()
    with ZipFile(input_buffer, "r") as source_zip:
        sheet_path = _sheet_path_for_name(source_zip, sheet_name)
        shared_strings = _shared_strings(source_zip)
        sheet_root = ET.fromstring(source_zip.read(sheet_path))
        headers = _headers_from_sheet(sheet_root, header_row, min_col, max_col, shared_strings)
        aligned = _align_to_pivot_headers(source_df, headers, dictionary)
        first_data_row = header_row + 1
        write_row = first_data_row if mode == "replace" else _last_row_in_columns(sheet_root, min_col, max_col) + 1
        new_last_row = write_row + len(aligned) - 1 if len(aligned) else write_row - 1
        old_last_row = max_row if mode == "replace" else write_row - 1
        new_ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{max(new_last_row, header_row)}"
        updated_sheet_xml = _replace_sheet_rows(
            sheet_root,
            aligned,
            min_col,
            max_col,
            first_data_row,
            old_last_row,
            write_row,
            mode,
        )

        with ZipFile(output_buffer, "w", ZIP_DEFLATED) as target_zip:
            for item in source_zip.infolist():
                data = source_zip.read(item.filename)
                if item.filename == sheet_path:
                    data = updated_sheet_xml
                elif item.filename == pivot_source.get("cache_definition"):
                    data = _set_pivot_source_ref(data, new_ref)
                target_zip.writestr(item, data)
    return output_buffer.getvalue()


def enable_pivot_refresh_on_load(workbook_bytes: bytes) -> bytes:
    input_buffer = BytesIO(workbook_bytes)
    output_buffer = BytesIO()
    with ZipFile(input_buffer, "r") as source_zip, ZipFile(output_buffer, "w", ZIP_DEFLATED) as target_zip:
        for item in source_zip.infolist():
            data = source_zip.read(item.filename)
            if item.filename.startswith("xl/pivotCache/") and item.filename.endswith(".xml"):
                try:
                    root = ET.fromstring(data)
                    root.set("refreshOnLoad", "1")
                    root.set("enableRefresh", "1")
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                except ET.ParseError:
                    pass
            target_zip.writestr(item, data)
    return output_buffer.getvalue()


def strip_pivot_cache_records(workbook_bytes: bytes) -> bytes:
    input_buffer = BytesIO(workbook_bytes)
    output_buffer = BytesIO()
    empty_records = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<pivotCacheRecords xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0"/>'
    )
    with ZipFile(input_buffer, "r") as source_zip, ZipFile(output_buffer, "w", ZIP_DEFLATED) as target_zip:
        for item in source_zip.infolist():
            data = source_zip.read(item.filename)
            if item.filename.startswith("xl/pivotCache/pivotCacheRecords") and item.filename.endswith(".xml"):
                data = empty_records
            elif item.filename.startswith("xl/pivotCache/") and item.filename.endswith(".xml"):
                data = _set_pivot_refresh_attributes(data)
            target_zip.writestr(item, data)
    return output_buffer.getvalue()


def _set_pivot_refresh_attributes(data: bytes) -> bytes:
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return data
    root.set("refreshOnLoad", "1")
    root.set("enableRefresh", "1")
    root.set("saveData", "0")
    if root.tag.endswith("pivotCacheDefinition"):
        root.set("recordCount", "0")
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _source_display(sheet: str | None, ref: str | None, source_name: str | None) -> str:
    if sheet and ref:
        return f"{sheet}!{ref}"
    if source_name:
        return source_name
    return "피벗 소스 범위 미확인"


def _effective_source_ref(archive: ZipFile, sheet_name: str | None, ref: str | None) -> str | None:
    if not sheet_name or not ref:
        return ref
    try:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
    except ValueError:
        return ref
    if max_row < 1000000:
        return ref
    try:
        sheet_path = _sheet_path_for_name(archive, sheet_name)
        sheet_root = ET.fromstring(archive.read(sheet_path))
    except Exception:
        return ref
    last_row = _last_row_in_columns(sheet_root, min_col, max_col)
    if last_row <= min_row:
        return ref
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{last_row}"


def _source_row_count(ref: str | None) -> int | None:
    if not ref:
        return None
    try:
        _, min_row, _, max_row = range_boundaries(ref)
    except ValueError:
        return None
    return max(0, max_row - min_row)


def _align_to_pivot_headers(
    source_df: pd.DataFrame,
    headers: list[Any],
    dictionary: StandardDictionary,
) -> pd.DataFrame:
    aligned = pd.DataFrame()
    matched_count = 0
    for header in headers:
        header_text = str(header).strip()
        standard_name = dictionary.match(header_text) or header_text
        if header_text in source_df.columns:
            aligned[header_text] = source_df[header_text]
            matched_count += 1
        elif standard_name in source_df.columns:
            aligned[header_text] = source_df[standard_name]
            matched_count += 1
        else:
            aligned[header_text] = None
    if headers and matched_count == 0:
        raise ValueError("피벗 소스 헤더와 원본 데이터 컬럼이 하나도 매칭되지 않았습니다.")
    return aligned


def _sheet_path_for_name(archive: ZipFile, sheet_name: str) -> str:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        rel.attrib.get("Id"): rel.attrib.get("Target")
        for rel in rels_root.findall("pkgrel:Relationship", NS)
    }
    for sheet in workbook_root.findall(".//main:sheet", NS):
        if sheet.attrib.get("name") != sheet_name:
            continue
        rel_id = sheet.attrib.get(f"{{{REL_NS}}}id")
        target = rel_targets.get(rel_id)
        if not target:
            break
        return _workbook_target_path(target)
    raise ValueError(f"피벗 소스 시트를 찾을 수 없습니다: {sheet_name}")


def _workbook_target_path(target: str) -> str:
    normalized = target.lstrip("/")
    if normalized.startswith("xl/"):
        return normalized
    return f"xl/{normalized}"


def _shared_strings(archive: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    strings = []
    for item in root.findall("main:si", NS):
        strings.append("".join(text.text or "" for text in item.findall(".//main:t", NS)))
    return strings


def _headers_from_sheet(
    sheet_root: ET.Element,
    header_row: int,
    min_col: int,
    max_col: int,
    shared_strings: list[str],
) -> list[str]:
    cells = _cells_by_column(sheet_root, header_row)
    headers = []
    for column in range(min_col, max_col + 1):
        cell = cells.get(column)
        headers.append(_cell_value(cell, shared_strings) if cell is not None else get_column_letter(column))
    return headers


def _cells_by_column(sheet_root: ET.Element, row_number: int) -> dict[int, ET.Element]:
    row = sheet_root.find(f".//main:row[@r='{row_number}']", NS)
    if row is None:
        return {}
    cells = {}
    for cell in row.findall("main:c", NS):
        ref = cell.attrib.get("r", "")
        column = _column_index(ref)
        if column:
            cells[column] = cell
    return cells


def _cell_value(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.findall(".//main:t", NS))
    value = cell.find("main:v", NS)
    if value is None or value.text is None:
        return ""
    if cell_type == "s":
        index = int(value.text)
        return shared_strings[index] if 0 <= index < len(shared_strings) else ""
    return value.text


def _replace_sheet_rows(
    sheet_root: ET.Element,
    aligned: pd.DataFrame,
    min_col: int,
    max_col: int,
    first_data_row: int,
    old_last_row: int,
    write_row: int,
    mode: str,
) -> bytes:
    sheet_data = sheet_root.find("main:sheetData", NS)
    if sheet_data is None:
        sheet_data = ET.SubElement(sheet_root, f"{{{MAIN_NS}}}sheetData")

    if mode == "replace":
        for row in list(sheet_data.findall("main:row", NS)):
            row_number = int(row.attrib.get("r", "0") or 0)
            if first_data_row <= row_number <= old_last_row:
                sheet_data.remove(row)

    for offset, row_values in enumerate(aligned.itertuples(index=False, name=None)):
        sheet_data.append(_row_element(write_row + offset, row_values, min_col, max_col))

    rows = sorted(sheet_data.findall("main:row", NS), key=lambda row: int(row.attrib.get("r", "0") or 0))
    for row in list(sheet_data.findall("main:row", NS)):
        sheet_data.remove(row)
    for row in rows:
        sheet_data.append(row)

    new_last_row = max(write_row + len(aligned) - 1, first_data_row - 1)
    _set_dimension(sheet_root, min_col, 1, max_col, new_last_row)
    return ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)


def _row_element(row_number: int, values: tuple[Any, ...], min_col: int, max_col: int) -> ET.Element:
    row = ET.Element(f"{{{MAIN_NS}}}row", {"r": str(row_number)})
    for offset, value in enumerate(values[: max_col - min_col + 1]):
        column = min_col + offset
        cell_ref = f"{get_column_letter(column)}{row_number}"
        row.append(_cell_element(cell_ref, _scalar(value)))
    return row


def _cell_element(cell_ref: str, value: Any) -> ET.Element:
    cell = ET.Element(f"{{{MAIN_NS}}}c", {"r": cell_ref})
    if value is None:
        return cell
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        value_node = ET.SubElement(cell, f"{{{MAIN_NS}}}v")
        value_node.text = str(value)
        return cell
    cell.set("t", "inlineStr")
    inline = ET.SubElement(cell, f"{{{MAIN_NS}}}is")
    text = ET.SubElement(inline, f"{{{MAIN_NS}}}t")
    text.text = str(value)
    return cell


def _last_row_in_columns(sheet_root: ET.Element, min_col: int, max_col: int) -> int:
    last_row = 1
    for row in sheet_root.findall(".//main:row", NS):
        row_number = int(row.attrib.get("r", "0") or 0)
        for cell in row.findall("main:c", NS):
            column = _column_index(cell.attrib.get("r", ""))
            if min_col <= column <= max_col and _has_cell_value(cell):
                last_row = max(last_row, row_number)
                break
    return last_row


def _has_cell_value(cell: ET.Element) -> bool:
    return cell.find("main:v", NS) is not None or cell.find("main:is", NS) is not None


def _column_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref.upper())
    if not match:
        return 0
    index = 0
    for char in match.group(1):
        index = index * 26 + ord(char) - 64
    return index


def _set_dimension(sheet_root: ET.Element, min_col: int, min_row: int, max_col: int, max_row: int) -> None:
    dimension = sheet_root.find("main:dimension", NS)
    if dimension is None:
        return
    dimension.set("ref", f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max(max_row, min_row)}")


def _set_pivot_source_ref(data: bytes, ref: str) -> bytes:
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return data
    worksheet_source = root.find(".//main:worksheetSource", NS)
    if worksheet_source is not None:
        worksheet_source.set("ref", ref)
    root.set("refreshOnLoad", "1")
    root.set("enableRefresh", "1")
    root.set("saveData", "0")
    root.set("recordCount", "0")
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _scalar(value: Any) -> Any:
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    return value
