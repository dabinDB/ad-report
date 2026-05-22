from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from .aggregation import aggregate_table, build_total_row
from .dictionary import StandardDictionary


def fill_workbook(
    template_bytes: bytes,
    source_df: pd.DataFrame,
    definitions: list[dict[str, Any]],
    dictionary: StandardDictionary,
) -> bytes:
    workbook = load_workbook(BytesIO(template_bytes))
    for definition in definitions:
        location = definition.get("location", {})
        sheet = workbook[location["sheet"]]
        result = aggregate_table(source_df, definition, dictionary)
        start_row = int(location["data_start_row"])
        end_row = int(location["data_end_row"])
        label_col = normalize_excel_column(location["label_col"])
        label_col_idx = column_index_from_string(label_col)
        metric_columns = normalize_metric_columns(location.get("columns", {}))
        group_by = definition.get("group_by", [])
        preserved_cells = formula_cells_to_preserve(definition)
        total_spec = normalize_summary_row(definition.get("total_row"), "합계")
        compare_spec = normalize_summary_row(definition.get("compare_row"), "전월 비교")
        average_spec = normalize_summary_row(definition.get("average_row"), "일평균")
        reserved_rows = {
            row
            for row in [total_spec.get("row"), average_spec.get("row"), compare_spec.get("row")]
            if isinstance(row, int) and start_row <= row <= end_row
        }

        _clear_range(sheet, start_row, end_row, [label_col, *metric_columns.keys()], preserved_cells)

        if total_spec.get("enabled") and total_spec.get("row"):
            total = build_total_row(source_df, definition, dictionary)
            _write_cell(sheet, total_spec["row"], label_col_idx, total_spec.get("label", "합계"), preserved_cells)
            _write_metric_cells(sheet, total_spec["row"], metric_columns, total, preserved_cells)

        if average_spec.get("enabled") and average_spec.get("row"):
            average = build_average_row(source_df, result, definition, dictionary)
            _write_cell(sheet, average_spec["row"], label_col_idx, average_spec.get("label", "일평균"), preserved_cells)
            _write_metric_cells(sheet, average_spec["row"], metric_columns, average, preserved_cells)

        if compare_spec.get("enabled") and compare_spec.get("row"):
            compare_values = build_compare_row(result, definition, dictionary)
            _write_cell(sheet, compare_spec["row"], label_col_idx, compare_spec.get("label", "전월 비교"), preserved_cells)
            _write_metric_cells(sheet, compare_spec["row"], metric_columns, compare_values, preserved_cells)

        write_row = start_row
        if total_spec.get("enabled") and not total_spec.get("row") and total_spec.get("position", "top") == "top":
            total = build_total_row(source_df, definition, dictionary)
            write_row = _next_data_row(write_row, reserved_rows)
            _write_cell(sheet, write_row, label_col_idx, total_spec.get("label", "합계"), preserved_cells)
            _write_metric_cells(sheet, write_row, metric_columns, total, preserved_cells)
            reserved_rows.add(write_row)
            write_row += 1

        for _, row in result.iterrows():
            write_row = _next_data_row(write_row, reserved_rows)
            if write_row > end_row:
                break
            _write_cell(sheet, write_row, label_col_idx, _label_value(row, group_by), preserved_cells)
            _write_metric_cells(sheet, write_row, metric_columns, row.to_dict(), preserved_cells)
            write_row += 1

        if total_spec.get("enabled") and not total_spec.get("row") and total_spec.get("position") == "bottom":
            write_row = _next_data_row(write_row, reserved_rows)
        if total_spec.get("enabled") and not total_spec.get("row") and total_spec.get("position") == "bottom" and write_row <= end_row:
            total = build_total_row(source_df, definition, dictionary)
            _write_cell(sheet, write_row, label_col_idx, total_spec.get("label", "합계"), preserved_cells)
            _write_metric_cells(sheet, write_row, metric_columns, total, preserved_cells)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _clear_range(sheet: Any, start_row: int, end_row: int, columns: list[str], preserved_cells: set[str]) -> None:
    for row in range(start_row, end_row + 1):
        for column in columns:
            column_idx = column_index_from_string(normalize_excel_column(column))
            _write_cell(sheet, row, column_idx, None, preserved_cells)


def _write_metric_cells(sheet: Any, row: int, metric_columns: dict[str, str], values: dict[str, Any], preserved_cells: set[str]) -> None:
    for column, metric in metric_columns.items():
        value = values.get(metric)
        _write_cell(sheet, row, column_index_from_string(normalize_excel_column(column)), _scalar(value), preserved_cells)


def _write_cell(sheet: Any, row: int, column_idx: int, value: Any, preserved_cells: set[str]) -> None:
    cell = sheet.cell(row, column_idx)
    if cell.coordinate in preserved_cells:
        return
    cell.value = value


def formula_cells_to_preserve(definition: dict[str, Any]) -> set[str]:
    if definition.get("formula_policy", "overwrite") != "preserve":
        return set()
    return {
        str(cell.get("cell"))
        for cell in definition.get("formula_cells", [])
        if cell.get("cell")
    }


def build_compare_row(result: pd.DataFrame, definition: dict[str, Any], dictionary: StandardDictionary) -> dict[str, Any]:
    metrics = definition.get("metrics", [])
    if len(result) < 2:
        return {metric: None for metric in metrics}
    current = result.iloc[0]
    previous = result.iloc[1]
    values: dict[str, Any] = {}
    for metric in metrics:
        current_value = _number(current.get(metric))
        previous_value = _number(previous.get(metric))
        diff = current_value - previous_value
        spec = dictionary.metric_spec(metric)
        if spec.get("unit") == "%" or spec.get("multiplier") == 100:
            values[metric] = f"{diff:,.2f}%P"
        else:
            change = None if previous_value == 0 else diff / previous_value * 100
            values[metric] = f"{diff:,.0f}" if change is None else f"{diff:,.0f}({change:,.2f}%)"
    return values


def build_average_row(
    source_df: pd.DataFrame,
    result: pd.DataFrame,
    definition: dict[str, Any],
    dictionary: StandardDictionary,
) -> dict[str, Any]:
    metrics = definition.get("metrics", [])
    day_count = _daily_denominator(source_df)
    total_values = build_total_row(source_df, definition, dictionary)
    values: dict[str, Any] = {}
    for metric in metrics:
        spec = dictionary.metric_spec(metric)
        if spec.get("aggregation") == "weighted_avg":
            values[metric] = total_values.get(metric)
        elif metric in source_df.columns and day_count:
            values[metric] = _number(source_df[metric].sum()) / day_count
        elif metric in result.columns and len(result):
            values[metric] = result[metric].apply(_number).mean()
        else:
            values[metric] = None
    return values


def normalize_summary_row(value: Any, default_label: str) -> dict[str, Any]:
    if isinstance(value, int):
        return {"enabled": True, "row": value, "label": default_label}
    if isinstance(value, dict):
        normalized = {
            "enabled": bool(value.get("enabled", value.get("row") is not None)),
            "label": value.get("label", default_label),
        }
        if value.get("row"):
            normalized["row"] = int(value["row"])
        if value.get("position"):
            normalized["position"] = value["position"]
        if value.get("mode"):
            normalized["mode"] = value["mode"]
        return normalized
    return {"enabled": False, "label": default_label}


def _daily_denominator(source_df: pd.DataFrame) -> int:
    if "날짜" not in source_df.columns:
        return 0
    dates = pd.to_datetime(source_df["날짜"], errors="coerce").dropna()
    return int(dates.dt.normalize().nunique())


def _next_data_row(row: int, reserved_rows: set[int]) -> int:
    while row in reserved_rows:
        row += 1
    return row


def normalize_metric_columns(columns: dict[Any, Any]) -> dict[str, str]:
    normalized = {}
    for column, metric in (columns or {}).items():
        try:
            normalized[normalize_excel_column(column)] = str(metric)
        except ValueError:
            normalized[normalize_excel_column(metric)] = str(column)
    return normalized


def normalize_excel_column(column: Any) -> str:
    if isinstance(column, int):
        if column < 1:
            raise ValueError(f"엑셀 컬럼 번호는 1 이상이어야 합니다: {column}")
        return get_column_letter(column)
    text = "" if column is None else str(column).strip().upper()
    if text.isdigit():
        return get_column_letter(int(text))
    if not text.isalpha():
        raise ValueError(f"엑셀 컬럼은 A, B, C 같은 문자여야 합니다: {column!r}")
    column_index_from_string(text)
    return text


def _label_value(row: pd.Series, group_by: list[str]) -> str:
    values = [str(row.get(column, "")) for column in group_by]
    return " / ".join(value for value in values if value and value != "nan")


def _scalar(value: Any) -> Any:
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    return value


def _number(value: Any) -> float:
    if pd.isna(value):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0
